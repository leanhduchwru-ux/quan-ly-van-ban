import streamlit as st
import pandas as pd
from database import get_connection
import uuid
import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.express as px

st.set_page_config(page_title="Hệ Thống Quản Lý Văn Bản", page_icon="📄", layout="wide")

st.markdown("""
<style>
    .stMetric {
        background-color: #1E293B;
        border-radius: 12px;
        padding: 20px;
        border: 1px solid rgba(255, 255, 255, 0.1);
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        transition: transform 0.3s ease;
    }
    .stMetric:hover { transform: translateY(-5px); border-color: rgba(255, 255, 255, 0.3); }
    h1 {
        background: -webkit-linear-gradient(45deg, #3b82f6, #8b5cf6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-weight: 700;
    }
</style>
""", unsafe_allow_html=True)

st.title("Hệ Thống Quản Lý Văn Bản Điều Hành")

# Tự động hiển thị Global Dashboard
def render_global_dashboard():
    conn = get_connection()
    try:
        total_in = conn.execute("SELECT COUNT(*) FROM documents WHERE type = 'INCOMING'").fetchone()[0]
        total_out = conn.execute("SELECT COUNT(*) FROM documents WHERE type = 'OUTGOING'").fetchone()[0]
        
        pending = conn.execute("SELECT COUNT(*) FROM document_relations WHERE match_status LIKE '%Chưa có%'").fetchone()[0]
        
        if total_in > 0 or total_out > 0:
            st.markdown("### 📊 Tổng Quan Toàn Cơ Quan")
            pending_class = "card-pending-alert" if pending > 0 else "card-pending-0"
            html_content = f"""
            <style>
            .metric-container {{
                display: flex;
                justify-content: space-between;
                gap: 15px;
                margin-bottom: 20px;
            }}
            .metric-card {{
                flex: 1;
                padding: 20px;
                border-radius: 10px;
                color: white;
                box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                font-family: sans-serif;
            }}
            .card-in {{ background: linear-gradient(135deg, #3b82f6, #2563eb); }}
            .card-out {{ background: linear-gradient(135deg, #10b981, #059669); }}
            .card-pending-0 {{ background: linear-gradient(135deg, #64748b, #475569); }}
            .card-pending-alert {{ 
                background: linear-gradient(135deg, #ef4444, #dc2626); 
                animation: pulse 2s infinite; 
            }}
            @keyframes pulse {{
                0% {{ box-shadow: 0 0 0 0 rgba(239, 68, 68, 0.7); }}
                70% {{ box-shadow: 0 0 0 10px rgba(239, 68, 68, 0); }}
                100% {{ box-shadow: 0 0 0 0 rgba(239, 68, 68, 0); }}
            }}
            .metric-title {{ font-size: 1.1rem; font-weight: 600; margin-bottom: 5px; opacity: 0.9; }}
            .metric-value {{ font-size: 2.5rem; font-weight: bold; margin: 0; }}
            </style>
            
            <div class="metric-container">
                <div class="metric-card card-in">
                    <div class="metric-title">📥 Tổng số Văn bản đến</div>
                    <div class="metric-value">{total_in}</div>
                </div>
                <div class="metric-card card-out">
                    <div class="metric-title">📤 Tổng số Văn bản đi</div>
                    <div class="metric-value">{total_out}</div>
                </div>
                <div class="metric-card {pending_class}">
                    <div class="metric-title">🔥 Tồn đọng cần xử lý</div>
                    <div class="metric-value">{pending}</div>
                </div>
            </div>
            """
            st.markdown(html_content, unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Row 1 of charts
            import plotly.express as px
            import pandas as pd
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Tỷ lệ trạng thái văn bản (Toàn Hệ thống)**")
                status_rows = conn.execute("SELECT match_status, COUNT(*) FROM document_relations GROUP BY match_status").fetchall()
                if status_rows:
                    status_df = pd.DataFrame(status_rows, columns=['Trạng thái', 'Số lượng'])
                    status_df['Trạng thái'] = status_df['Trạng thái'].apply(lambda x: 
                        '🟡 Đang nhận việc' if 'Đang' in x else
                        '🔴 Chưa có trả lời' if 'Chưa' in x else
                        '🟢 Có văn bản đi' if 'Có' in x else
                        '🔵 Nhận để biết' if 'biết' in x else x
                    )
                    fig_status = px.pie(status_df, values='Số lượng', names='Trạng thái', hole=0.4,
                         color='Trạng thái',
                         color_discrete_map={
                             '🟡 Đang nhận việc': '#eab308', 
                             '🔴 Chưa có trả lời': '#ef4444', 
                             '🟢 Có văn bản đi': '#22c55e',
                             '🔵 Nhận để biết': '#3b82f6'
                         })
                    fig_status.update_layout(margin=dict(t=20, b=20, l=0, r=0), showlegend=True, legend=dict(orientation="h", y=-0.2))
                    st.plotly_chart(fig_status, use_container_width=True)
            
            with c2:
                st.markdown("**Nguồn phân bổ văn bản**")
                source_rows = conn.execute("SELECT content, type, COUNT(*) FROM documents GROUP BY content, type").fetchall()
                if source_rows:
                    source_df = pd.DataFrame(source_rows, columns=['Hệ thống', 'Loại', 'Số lượng'])
                    source_df['Phân loại'] = source_df.apply(lambda row: f"{row['Hệ thống']} {'Đến' if row['Loại'] == 'INCOMING' else 'Đi'}", axis=1)
                    fig_source = px.pie(source_df, values='Số lượng', names='Phân loại', hole=0.6,
                        color_discrete_sequence=px.colors.sequential.Teal)
                    fig_source.update_layout(margin=dict(t=20, b=20, l=0, r=0), showlegend=True, legend=dict(orientation="h", y=-0.2))
                    st.plotly_chart(fig_source, use_container_width=True)

            st.markdown("<br>", unsafe_allow_html=True)
            
            # Row 2 of charts
            c3, c4 = st.columns(2)
            with c3:
                st.markdown("**Top 10 cá nhân/đơn vị nợ việc nhất (Cả 2 HT)**")
                unanswered_tasks = conn.execute('''
                    SELECT t.assignee FROM tasks t 
                    JOIN document_relations r ON t.document_id = r.incoming_id 
                    WHERE r.match_status LIKE '%Chưa có%'
                ''').fetchall()
                if unanswered_tasks:
                    assignee_counts = {}
                    for row in unanswered_tasks:
                        if row[0]:
                            for person in str(row[0]).split(','):
                                p = person.strip()
                                if p:
                                    assignee_counts[p] = assignee_counts.get(p, 0) + 1
                    if assignee_counts:
                        bar_data = pd.DataFrame(list(assignee_counts.items()), columns=['Người xử lý', 'Số lượng'])
                        bar_data = bar_data.sort_values(by='Số lượng', ascending=False).head(10)
                        fig_bar = px.bar(bar_data, x='Số lượng', y='Người xử lý', orientation='h', color_discrete_sequence=['#ef4444'])
                        fig_bar.update_layout(yaxis={'categoryorder':'total ascending'}, margin=dict(t=20, b=20, l=0, r=0))
                        st.plotly_chart(fig_bar, use_container_width=True)
                    else:
                        st.success("Tuyệt vời! Không có cá nhân nào bị tồn đọng việc.")
                else:
                    st.success("Tuyệt vời! Cơ quan không có văn bản nợ đọng.")
            
            with c4:
                st.markdown("**So sánh tiến độ giữa VOFFICE và HPNET**")
                sys_status_rows = conn.execute('''
                    SELECT i.content, r.match_status, COUNT(*) 
                    FROM document_relations r
                    JOIN documents i ON r.incoming_id = i.id
                    GROUP BY i.content, r.match_status
                ''').fetchall()
                
                if sys_status_rows:
                    sys_df = pd.DataFrame(sys_status_rows, columns=['Hệ thống', 'Trạng thái gốc', 'Số lượng'])
                    sys_df['Trạng thái'] = sys_df['Trạng thái gốc'].apply(lambda x: 
                        'Đang nhận' if 'Đang' in x else
                        'Chưa trả lời' if 'Chưa' in x else
                        'Có VB đi' if 'Có' in x else
                        'Để biết' if 'biết' in x else x
                    )
                    fig_stack = px.bar(sys_df, x="Hệ thống", y="Số lượng", color="Trạng thái",
                        color_discrete_map={
                            'Đang nhận': '#eab308', 
                            'Chưa trả lời': '#ef4444', 
                            'Có VB đi': '#22c55e',
                            'Để biết': '#3b82f6'
                        }, barmode='stack')
                    fig_stack.update_layout(margin=dict(t=20, b=20, l=0, r=0), legend=dict(orientation="h", y=-0.2))
                    st.plotly_chart(fig_stack, use_container_width=True)
            # Phân tích Tồn đọng chuyên sâu
            st.markdown("### 🕵️‍♂️ Phân Tích Tồn Đọng Chuyên Sâu")
            
            unanswered_docs = conn.execute('''
                SELECT 
                    i.content as system_name,
                    i.document_no,
                    i.summary,
                    i.issued_date,
                    t.assignee
                FROM document_relations r
                JOIN documents i ON r.incoming_id = i.id
                JOIN tasks t ON r.incoming_id = t.document_id
                WHERE r.match_status LIKE '%Chưa có%'
            ''').fetchall()

            if unanswered_docs:
                pending_df = pd.DataFrame(unanswered_docs, columns=['Hệ thống', 'Ký hiệu', 'Trích yếu', 'Ngày đến', 'Người xử lý'])
                
                assignees_set = set()
                for a in pending_df['Người xử lý']:
                    if a:
                        assignees_set.update([x.strip() for x in str(a).split(',') if x.strip()])
                
                f_col1, f_col2 = st.columns([1, 2])
                with f_col1:
                    selected_assignee = st.selectbox("Tra cứu văn bản nợ đọng theo Cá nhân/Đơn vị:", ["Tất cả"] + sorted(list(assignees_set)))
                
                filtered_pending = pending_df.copy()
                if selected_assignee != "Tất cả":
                    filtered_pending = filtered_pending[filtered_pending['Người xử lý'].str.contains(selected_assignee, na=False)]
                
                if not filtered_pending.empty:
                    filtered_pending.insert(0, 'TT', range(1, 1 + len(filtered_pending)))
                    
                    def color_pending(val):
                        return 'color: #ef4444; font-weight: bold;'
                        
                    st.dataframe(
                        filtered_pending.style.map(color_pending, subset=['Ký hiệu']),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "TT": st.column_config.NumberColumn("TT", width="small"),
                            "Hệ thống": st.column_config.TextColumn("Hệ thống", width="small"),
                            "Ký hiệu": st.column_config.TextColumn("Ký hiệu", width="medium"),
                            "Trích yếu": st.column_config.TextColumn("Trích yếu", width="large"),
                            "Ngày đến": st.column_config.TextColumn("Ngày đến", width="small"),
                            "Người xử lý": st.column_config.TextColumn("Người xử lý", width="medium")
                        }
                    )
                else:
                    st.info("Không có văn bản nợ đọng cho cá nhân/đơn vị này.")
                    
            st.markdown("---")
    finally:
        conn.close()

render_global_dashboard()

def match_documents(system_name=None):
    conn = get_connection()
    try:
        if system_name:
            incoming_docs = conn.execute("SELECT * FROM documents WHERE type = 'INCOMING' AND content = ?", (system_name,)).fetchall()
            outgoing_docs = conn.execute("SELECT * FROM documents WHERE type = 'OUTGOING' AND content = ?", (system_name,)).fetchall()
        else:
            incoming_docs = conn.execute("SELECT * FROM documents WHERE type = 'INCOMING'").fetchall()
            outgoing_docs = conn.execute("SELECT * FROM documents WHERE type = 'OUTGOING'").fetchall()
        
        updated_count = 0
        for incoming in incoming_docs:
            current_status = "Đang nhận việc"
            matched_out_id = None
            
            # Match
            def is_matched(inc, out):
                inc_no = str(inc['document_no']).strip().lower()
                out_sum = str(out['summary']).strip().lower() if out['summary'] else ""
                inc_sum = str(inc['summary']).strip().lower() if inc['summary'] else ""
                
                if inc_no and inc_no != 'nan' and inc_no in out_sum:
                    return True
                
                if inc_sum and out_sum:
                    if inc_sum == out_sum or inc_sum in out_sum or out_sum in inc_sum:
                        return True
                        
                return False

            matched_out = next((out for out in outgoing_docs if is_matched(incoming, out)), None)
            
            if matched_out:
                current_status = "Có văn bản đi"
                matched_out_id = matched_out['id']
            else:
                summary_lower = str(incoming['summary']).lower() if incoming['summary'] else ""
                exclusion_keywords = ['thông báo', 'giấy mời', 'tuyên truyền', 'phổ biến', 'tin buồn', 'họp', 'để biết', 'gửi tài liệu', 'triệu tập', 'quán triệt']
                
                is_info_only = any(kw in summary_lower for kw in exclusion_keywords)
                
                if is_info_only:
                    current_status = "Nhận để biết (Không cần trả lời)"
                else:
                    is_overdue = incoming['deadline'] and datetime.datetime.now() > datetime.datetime.fromisoformat(incoming['deadline'])
                    tasks = conn.execute("SELECT * FROM tasks WHERE document_id = ?", (incoming['id'],)).fetchall()
                    has_pending = any(t['status'] != 'Hoàn thành' for t in tasks)
                    
                    if is_overdue or (not has_pending and len(tasks) > 0):
                        current_status = "Chưa có văn bản trả lời"
                    else:
                        current_status = "Đang nhận việc"
            
            relation = conn.execute("SELECT id FROM document_relations WHERE incoming_id = ?", (incoming['id'],)).fetchone()
            if not relation:
                conn.execute("INSERT INTO document_relations (id, incoming_id, outgoing_id, match_status) VALUES (?, ?, ?, ?)",
                             (str(uuid.uuid4()), incoming['id'], matched_out_id, current_status))
            else:
                conn.execute("UPDATE document_relations SET outgoing_id = ?, match_status = ? WHERE id = ?",
                             (matched_out_id, current_status, relation['id']))
            updated_count += 1
            
        conn.commit()
        return updated_count
    finally:
        conn.close()

def generate_mock_data():
    conn = get_connection()
    try:
        # Clear old data
        conn.execute("DELETE FROM document_relations")
        conn.execute("DELETE FROM tasks")
        conn.execute("DELETE FROM documents")
        
        # Insert Incoming
        conn.execute("INSERT INTO documents (id, type, document_no, summary, deadline) VALUES ('in1', 'INCOMING', '123/UBND', 'V/v báo cáo tiến độ dự án', '2027-12-31 23:59:59')")
        conn.execute("INSERT INTO documents (id, type, document_no, summary, deadline) VALUES ('in2', 'INCOMING', '456/STT', 'Yêu cầu phúc đáp công văn', '2022-01-01 23:59:59')")
        conn.execute("INSERT INTO documents (id, type, document_no, summary, deadline) VALUES ('in3', 'INCOMING', '789/VP', 'Xin ý kiến chỉ đạo của Giám đốc', '2027-12-31 23:59:59')")
        
        # Insert Outgoing
        conn.execute("INSERT INTO documents (id, type, document_no, summary) VALUES ('out1', 'OUTGOING', '102/UBND-TL', 'Trả lời công văn 789/VP')")
        
        # Insert Tasks
        conn.execute("INSERT INTO tasks (id, document_id, assignee, status) VALUES ('t1', 'in1', 'Nguyễn Văn A', 'Đang xử lý')")
        conn.execute("INSERT INTO tasks (id, document_id, assignee, status) VALUES ('t2', 'in3', 'Trần Thị B', 'Hoàn thành')")
        
        conn.commit()
    finally:
        conn.close()

st.markdown("### 📥 Nạp dữ liệu từ File Excel")
st.markdown("Vui lòng tải lên file danh sách xuất ra từ hệ thống. File cần có các cột chứa chữ **'Ký hiệu'** và **'Trích yếu'**.")

col_up1, col_up2 = st.columns(2)

def process_uploaded_file(file, doc_type, system_name):
    if file is not None:
        try:
            df = pd.read_excel(file, header=None)
            conn = get_connection()
            
            # Xóa dữ liệu cũ của riêng Hệ thống này (VOFFICE hoặc HPNET) để tránh trùng lặp
            # Lưu ý: Cột content được dùng để lưu tên hệ thống (VOFFICE/HPNET)
            conn.execute("DELETE FROM documents WHERE type = ? AND content = ?", (doc_type, system_name))
            
            # Lấy danh sách văn bản đã có để tránh trùng lặp (Kết hợp cả Ký hiệu và Trích yếu)
            existing_docs = set()
            for r in conn.execute("SELECT document_no, summary FROM documents WHERE type = ?", (doc_type,)).fetchall():
                key = f"{str(r['document_no']).strip().lower()}_{str(r['summary']).strip().lower()}"
                existing_docs.add(key)
            
            count = 0
            doc_col = -1
            summary_col = -1
            assignee_col = -1
            date_col = -1
            agency_col = -1
            
            import re
            
            for index, row in df.iterrows():
                # Bươc 1: Dò tìm dòng chứa Tiêu đề cột
                if doc_col == -1 or summary_col == -1:
                    for i, val in enumerate(row):
                        if pd.notna(val):
                            val_str = str(val).lower()
                            if 'ký hiệu' in val_str or 's.k.h' in val_str:
                                doc_col = i
                            if 'trích yếu' in val_str:
                                summary_col = i
                            if 'xử lý' in val_str or 'người nhận' in val_str or 'chuyển' in val_str:
                                assignee_col = i
                            if 'ngày ban hành' in val_str or 'ngày nhận' in val_str or 'ngày văn bản' in val_str or 'ngày' in val_str.split():
                                date_col = i
                            if 'nơi ban hành' in val_str or 'nơi lưu trữ' in val_str or 'nơi nhận' in val_str or 'cơ quan' in val_str:
                                agency_col = i
                    continue
                
                # Bước 2: Đọc dữ liệu từ các dòng bên dưới
                if doc_col != -1 and summary_col != -1:
                    doc_number = str(row[doc_col]).strip() if pd.notna(row[doc_col]) else ""
                    summary = str(row[summary_col]).strip() if pd.notna(row[summary_col]) else ""
                    
                    doc_date = ""
                    if date_col != -1 and pd.notna(row[date_col]):
                        val = row[date_col]
                        if isinstance(val, datetime.datetime):
                            doc_date = val.strftime("%d/%m/%Y")
                        else:
                            doc_date = str(val).strip()
                            # Loại bỏ "00:00:00" nếu pandas ép kiểu nhầm thành string có giờ
                            doc_date = doc_date.replace("00:00:00", "").strip()
                            
                    agency = str(row[agency_col]).strip() if agency_col != -1 and pd.notna(row[agency_col]) else ""
                    
                    if doc_number and doc_number.lower() not in ['nan', 'none', '']:
                        key_to_check = f"{doc_number.strip().lower()}_{summary.strip().lower()}"
                        if key_to_check not in existing_docs:
                            doc_id = str(uuid.uuid4())
                            # Lưu tên hệ thống vào cột content, lưu cơ quan vào cột system_source
                            conn.execute("INSERT INTO documents (id, type, document_no, issued_date, system_source, summary, content) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                         (doc_id, doc_type, doc_number, doc_date, agency, summary, system_name))
                            
                            # Xử lý bóc tách Người xử lý (Lấy chính xác người liền sau Trương Mạnh Tiến)
                            if assignee_col != -1 and pd.notna(row[assignee_col]):
                                val = str(row[assignee_col]).strip()
                                parts = [p.strip() for p in re.split(r'[,\n;]', val) if p.strip()]
                                assignee = ""
                                
                                tm_index = -1
                                for i, p in enumerate(parts):
                                    if 'Trương Mạnh Tiến' in p:
                                        tm_index = i
                                        break
                                
                                if tm_index != -1 and tm_index < len(parts) - 1:
                                    # Lấy người liền sau Trương Mạnh Tiến
                                    assignee = parts[tm_index + 1]
                                elif tm_index == -1 and parts:
                                    # Không có Trương Mạnh Tiến, bỏ qua từ khóa 'Lưu'
                                    filtered = [p for p in parts if p.lower() not in ['lưu', 'lưu trữ', 'văn thư', 'vt']]
                                    assignee = filtered[0] if filtered else parts[0]
                                else:
                                    assignee = "Trương Mạnh Tiến"
                                
                                # Xóa bớt chữ thừa như "(Xử lý chính)", "(Phối hợp)" để tên gọn gàng
                                assignee = re.sub(r'\(.*?\)', '', assignee).strip()
                                
                                conn.execute("INSERT INTO tasks (id, document_id, assignee, status) VALUES (?, ?, ?, ?)",
                                             (str(uuid.uuid4()), doc_id, assignee, 'Đang xử lý'))
                            
                            existing_docs.add(key_to_check)
                            count += 1
            
            if doc_col == -1 or summary_col == -1:
                st.error("File không đúng định dạng cấu trúc, vui lòng kiểm tra lại cột Ký hiệu hoặc Trích yếu")
                conn.close()
                return 0

            conn.commit()
            conn.close()
            return count
        except Exception as e:
            st.error(f"Lỗi đọc file: {e}")
            return 0
    return 0

with col_up1:
    st.info("📥 VĂN BẢN ĐẾN")
    file_in_voffice = st.file_uploader("Kéo thả file Văn bản đến - VOFFICE", type=["xlsx", "xls"], key="in_voffice")
    if file_in_voffice:
        with st.spinner("Đang xử lý VOFFICE..."):
            c = process_uploaded_file(file_in_voffice, 'INCOMING', 'VOFFICE')
            match_documents('VOFFICE')
            st.success(f"Đã nạp {c} văn bản đến VOFFICE và tự động đối chiếu!")
            
    file_in_hpnet = st.file_uploader("Kéo thả file Văn bản đến - HPNET", type=["xlsx", "xls"], key="in_hpnet")
    if file_in_hpnet:
        with st.spinner("Đang xử lý HPNET..."):
            c = process_uploaded_file(file_in_hpnet, 'INCOMING', 'HPNET')
            match_documents('HPNET')
            st.success(f"Đã nạp {c} văn bản đến HPNET và tự động đối chiếu!")

with col_up2:
    st.info("📤 VĂN BẢN ĐI")
    file_out_voffice = st.file_uploader("Kéo thả file Văn bản đi - VOFFICE", type=["xlsx", "xls"], key="out_voffice")
    if file_out_voffice:
        with st.spinner("Đang xử lý VOFFICE..."):
            c = process_uploaded_file(file_out_voffice, 'OUTGOING', 'VOFFICE')
            match_documents('VOFFICE')
            st.success(f"Đã nạp {c} văn bản đi VOFFICE và tự động đối chiếu!")
            
    file_out_hpnet = st.file_uploader("Kéo thả file Văn bản đi - HPNET", type=["xlsx", "xls"], key="out_hpnet")
    if file_out_hpnet:
        with st.spinner("Đang xử lý HPNET..."):
            c = process_uploaded_file(file_out_hpnet, 'OUTGOING', 'HPNET')
            match_documents('HPNET')
            st.success(f"Đã nạp {c} văn bản đi HPNET và tự động đối chiếu!")

if st.button("🔄 Chạy Đối Chiếu Tự Động TOÀN HỆ THỐNG (VOFFICE + HPNET)", type="primary", use_container_width=True):
    with st.spinner("Đang phân tích và đối chiếu toàn bộ hệ thống..."):
        up_voffice = match_documents('VOFFICE')
        up_hpnet = match_documents('HPNET')
        st.success(f"Đã đối chiếu thành công {up_voffice} văn bản VOFFICE và {up_hpnet} văn bản HPNET! Vui lòng xem kết quả bên dưới.")


st.markdown("---")

conn = get_connection()
relations = conn.execute('''
    SELECT 
        r.match_status,
        i.document_no as inc_doc,
        i.issued_date as inc_date,
        i.system_source as inc_agency,
        i.summary as inc_summary,
        i.content as system_name,
        o.document_no as out_doc,
        o.issued_date as out_date,
        o.system_source as out_agency,
        i.id as inc_id
    FROM document_relations r
    JOIN documents i ON r.incoming_id = i.id
    LEFT JOIN documents o ON r.outgoing_id = o.id
''').fetchall()

data_voffice = []
data_hpnet = []
stats_voffice = {"Đang nhận việc": 0, "Chưa có văn bản trả lời": 0, "Có văn bản đi": 0, "Nhận để biết (Không cần trả lời)": 0}
data_hpnet = []
stats_hpnet = {"Đang nhận việc": 0, "Chưa có văn bản trả lời": 0, "Có văn bản đi": 0, "Nhận để biết (Không cần trả lời)": 0}

if len(relations) == 0:
    doc_count = conn.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
    if doc_count > 0:
        st.warning("Dữ liệu đã được nạp! Vui lòng bấm nút **'🔄 Chạy Đối Chiếu Tự Động'** ở phía trên để hệ thống bắt đầu phân tích và hiển thị bảng.")
    else:
        st.info("Cơ sở dữ liệu đang trống. Vui lòng tải dữ liệu từ file Excel lên các ô bên trên.")
else:
    for rel in relations:
        sys_name = rel['system_name']
        status = rel['match_status']
        
        tasks = conn.execute("SELECT assignee FROM tasks WHERE document_id = ?", (rel['inc_id'],)).fetchall()
        assignees = ", ".join([t['assignee'] for t in tasks if t['assignee']]) if tasks else ""
        
        display_status = status
        if status == 'Đang nhận việc':
            display_status = '🟡 Đang nhận việc'
        elif status == 'Chưa có văn bản trả lời':
            display_status = '🔴 Chưa có trả lời'
        elif status == 'Có văn bản đi':
            display_status = '🟢 Có văn bản đi'
            
        row_data = {
            "Văn bản đến": rel['inc_doc'],
            "ngày đến": rel['inc_date'] if rel['inc_date'] else "",
            "Nơi gửi đến": rel['inc_agency'] if rel['inc_agency'] else "",
            "Trích yếu": rel['inc_summary'],
            "Người xử lý": assignees,
            "VB đi": rel['out_doc'] if rel['out_doc'] else "",
            "Ngày gửi đi": rel['out_date'] if rel['out_date'] else "",
            "Nơi gửi đi": rel['out_agency'] if rel['out_agency'] else "",
            "Trạng thái": display_status
        }
        
        if sys_name == 'VOFFICE':
            stats_voffice[status] = stats_voffice.get(status, 0) + 1
            data_voffice.append(row_data)
        elif sys_name == 'HPNET':
            stats_hpnet[status] = stats_hpnet.get(status, 0) + 1
            data_hpnet.append(row_data)
        else:
            # Dự phòng nếu không có tên hệ thống
            stats_hpnet[status] = stats_hpnet.get(status, 0) + 1
            data_hpnet.append(row_data)

conn.close()

def color_status(val):
    if 'Đang nhận việc' in str(val): return 'color: #eab308; font-weight: bold;'
    elif 'Chưa có' in str(val): return 'color: #ef4444; font-weight: bold;'
    elif 'Có văn bản đi' in str(val): return 'color: #22c55e; font-weight: bold;'
    elif 'Nhận để biết' in str(val): return 'color: #3b82f6; font-weight: bold;'
    return ''

def generate_excel_report(df, system_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo Cáo"
    
    # Quốc hiệu tiêu ngữ
    ws.merge_cells('A1:J1')
    ws['A1'] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"
    ws['A1'].font = Font(name='Times New Roman', size=13, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:J2')
    ws['A2'] = "Độc lập - Tự do - Hạnh phúc"
    ws['A2'].font = Font(name='Times New Roman', size=14, bold=True, underline='single')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A4:J4')
    ws['A4'] = f"BẢNG KÊ ĐỐI CHIẾU VĂN BẢN ĐẾN - {system_name}"
    ws['A4'].font = Font(name='Times New Roman', size=14, bold=True)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = Font(name='Times New Roman', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = header_fill
        
    # Data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 7):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name='Times New Roman', size=12)
            cell.border = thin_border
            
            # Wrap text for Trích yếu (cột 5)
            if c_idx == 5:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top')
                
            # Color status (cột 10)
            if c_idx == 10 and value:
                if 'Đang nhận việc' in str(value):
                    cell.font = Font(name='Times New Roman', size=12, color='CA8A04', bold=True)
                elif 'Chưa có' in str(value):
                    cell.font = Font(name='Times New Roman', size=12, color='DC2626', bold=True)
                elif 'Có văn bản đi' in str(value):
                    cell.font = Font(name='Times New Roman', size=12, color='16A34A', bold=True)
                elif 'Nhận để biết' in str(value):
                    cell.font = Font(name='Times New Roman', size=12, color='2563EB', bold=True)

    # Column widths
    widths = [5, 15, 12, 20, 40, 15, 15, 12, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# Bổ sung nút tải báo cáo ngay dưới dữ liệu
st.markdown("### 📥 TẢI BÁO CÁO CHI TIẾT")
st.info("Sử dụng các nút bên dưới để tải bảng kê chi tiết toàn bộ văn bản của từng hệ thống ra file Excel.")
btn_col1, btn_col2 = st.columns(2)

with btn_col1:
    if data_voffice:
        df_voffice = pd.DataFrame(data_voffice)
        df_voffice.insert(0, 'TT', range(1, 1 + len(df_voffice)))
        excel_data_voffice = generate_excel_report(df_voffice, "VOFFICE")
        st.download_button(
            label=f"📥 TẢI XUỐNG BÁO CÁO VOFFICE",
            data=excel_data_voffice,
            file_name=f"BaoCao_DoiChieu_VOFFICE.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_VOFFICE",
            use_container_width=True
        )
        
with btn_col2:
    if data_hpnet:
        df_hpnet = pd.DataFrame(data_hpnet)
        df_hpnet.insert(0, 'TT', range(1, 1 + len(df_hpnet)))
        excel_data_hpnet = generate_excel_report(df_hpnet, "HPNET")
        st.download_button(
            label=f"📥 TẢI XUỐNG BÁO CÁO HPNET",
            data=excel_data_hpnet,
            file_name=f"BaoCao_DoiChieu_HPNET.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_HPNET",
            use_container_width=True
        )

st.markdown("---")
st.markdown("### 🐙 ĐỒNG BỘ GIAO VIỆC LÊN GITHUB KANBAN")
st.info("Tự động hóa luồng giao việc: Hệ thống sẽ tự động quét danh sách Tồn đọng và khởi tạo các thẻ công việc (Issues) tương ứng trên bảng Kanban của GitHub để đôn đốc.")

gh_col1, gh_col2 = st.columns([3, 1])
with gh_col1:
    gh_token = st.text_input("Nhập mã Personal Access Token (PAT) của GitHub:", type="password", help="Mã Token bắt đầu bằng 'ghp_'. Cần có quyền 'repo' và 'project'.")

with gh_col2:
    st.markdown("<br>", unsafe_allow_html=True) # Spacer
    if st.button("🚀 Đẩy Tồn đọng lên GitHub", use_container_width=True, type="primary"):
        if not gh_token:
            st.error("Vui lòng nhập mã Token!")
        else:
            with st.spinner("Đang kết nối GitHub và tạo thẻ công việc..."):
                try:
                    import github_sync
                    success, msg = github_sync.sync_to_github(token=gh_token)
                    if success:
                        st.success(msg)
                        st.balloons()
                    else:
                        st.error(msg)
                except Exception as e:
                    st.error(f"Lỗi hệ thống khi chạy kịch bản: {str(e)}")
